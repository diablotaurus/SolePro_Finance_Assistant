"""
Tests for transaction export service.
"""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from uuid import uuid4

import openpyxl
import pandas as pd

from solepro.core.application.dto.transaction_dto import TransactionResponseDTO
from solepro.presentation.desktop.services.transaction_export_service import (
    TransactionExportService,
)


def _tx(
    *,
    date: datetime,
    income: str,
    expense: str,
    tax: str,
    counterparty_name: str | None = None,
    note: str | None = None,
) -> TransactionResponseDTO:
    income_dec = Decimal(income)
    expense_dec = Decimal(expense)
    tax_dec = Decimal(tax)
    return TransactionResponseDTO(
        id=uuid4(),
        date=date,
        income=income_dec,
        expense=expense_dec,
        tax=tax_dec,
        profit=income_dec - expense_dec - tax_dec,
        counterparty_id=None,
        counterparty_name=counterparty_name,
        note=note,
        created_at=date,
        updated_at=date,
    )


def test_export_to_excel_writes_expected_columns_and_rows(tmp_path):
    service = TransactionExportService()
    output_path = tmp_path / "transactions.xlsx"

    rows = [
        _tx(
            date=datetime(2026, 2, 20, 11, 30),
            income="100.50",
            expense="20.25",
            tax="5.00",
            counterparty_name="Alpha",
            note="A",
        ),
        _tx(
            date=datetime(2026, 2, 21, 9, 15),
            income="0",
            expense="10",
            tax="0",
            counterparty_name=None,
            note=None,
        ),
    ]

    service.export_to_excel(rows, str(output_path))

    assert output_path.exists()
    with pd.ExcelFile(output_path) as xls:
        df = pd.read_excel(xls, sheet_name="Транзакции")
    assert list(df.columns) == [
        "Дата",
        "Доход",
        "Расход",
        "Налог",
        "Прибыль",
        "Контрагент",
        "Примечание",
    ]
    assert len(df.index) == 2
    assert df.loc[0, "Дата"] == "20.02.2026"
    assert float(df.loc[0, "Доход"]) == 100.5
    assert pd.isna(df.loc[1, "Контрагент"])
    assert pd.isna(df.loc[1, "Примечание"])


def test_export_to_excel_handles_empty_transactions(tmp_path):
    service = TransactionExportService()
    output_path = tmp_path / "empty.xlsx"

    service.export_to_excel([], str(output_path))

    assert output_path.exists()
    workbook = openpyxl.load_workbook(output_path)
    try:
        assert "Транзакции" in workbook.sheetnames
    finally:
        workbook.close()


def test_export_to_excel_limits_column_width(tmp_path):
    service = TransactionExportService()
    output_path = tmp_path / "wide.xlsx"
    very_long_note = "x" * 500

    service.export_to_excel(
        [
            _tx(
                date=datetime(2026, 2, 21, 9, 0),
                income="1000",
                expense="1",
                tax="1",
                counterparty_name="Name",
                note=very_long_note,
            )
        ],
        str(output_path),
    )

    workbook = openpyxl.load_workbook(output_path)
    try:
        sheet = workbook["Транзакции"]
        # Column G = "Примечание"
        assert sheet.column_dimensions["G"].width <= 50
    finally:
        workbook.close()
